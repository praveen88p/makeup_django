import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
from tkinter import Tk, filedialog, messagebox, simpledialog, Toplevel, Label, Button, StringVar, ttk
import os
from openpyxl.styles import Font  


# Constants
SEAT_POSITIONS = ["Left", "Middle", "Right"]
BORDER_STYLE = Border(left=Side(style='medium'), right=Side(style='medium'),
                      top=Side(style='medium'), bottom=Side(style='medium'))

# Initialize a dictionary to track the last used roll numbers for each position
last_used_roll_numbers = {"Left": "", "Middle": "", "Right": ""}

def prompt_for_new_roll_numbers(position):
    """Prompt the user to manually select the file for additional roll numbers using a file dialog."""
    root = Tk()
    root.withdraw()  # Hide the root window
    messagebox.showinfo("Roll Numbers Exhausted", f"Please select a new file for {position} roll numbers.")
    
    input_path = filedialog.askopenfilename(title=f"Select {position} roll numbers file", filetypes=[("Excel files", "*.xlsx *.xls")])
    
    if os.path.exists(input_path):
        print(f"Selected file for {position} roll numbers: {input_path}")
        return load_roll_numbers_from_file(input_path, position)  # Function to read the roll numbers
    else:
        print(f"Invalid file path for {position}.")
        return None


def load_roll_numbers_from_file(filepath, position):
    """Load roll numbers from the specified Excel file."""
    try:
        df = pd.read_excel(filepath)
        df.columns = df.columns.str.strip()

        if 'Roll Number' not in df.columns:
            messagebox.showerror("Error", f"'Roll Number' column not found in the Excel file for {position} roll numbers.")
            return None

        return df['Roll Number'].dropna().tolist()
    except Exception as e:
        messagebox.showerror("Error", f"Failed to read roll numbers file for {position} roll numbers: {e}")
        return None


def generate_seating_chart(room_number, rows, benches_per_row, students_per_bench, roll_numbers_lists, roll_number_indices, progress_var):
    seating_chart = []
    roll_numbers_exhausted = False
    total_seats = rows * benches_per_row * students_per_bench
    seat_counter = 0

    for row in range(rows):
        for bench in range(benches_per_row):
            for seat in range(students_per_bench):
                seat_number = seat + 1
                roll_numbers = roll_numbers_lists[seat]
                roll_number_index = roll_number_indices[seat]
                if roll_number_index < len(roll_numbers):
                    roll_number = roll_numbers[roll_number_index]
                    roll_number_indices[seat] += 1
                else:
                    roll_number = ''
                    roll_numbers_exhausted = True
                
                seating_chart.append([room_number, f'Row {row+1}', f'Bench {bench+1}', seat_number, roll_number])

                # Update the progress bar
                seat_counter += 1
                progress_var.set(int((seat_counter / total_seats) * 100))

    columns = ['Room Number', 'Row', 'Bench', 'Seat', 'Roll Number']
    df = pd.DataFrame(seating_chart, columns=columns)
    return df, roll_number_indices, roll_numbers_exhausted


def save_to_excel(seating_chart_list, roll_ranges_list, output_filename):
    """Save the seating chart and roll ranges to an Excel file."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Seating Chart"

    for seating_df in seating_chart_list:
        room_number = seating_df['Room Number'].iloc[0]

        # Add "Room Number" above the rows, centered
        add_room_header(ws1, room_number, seating_df)

        # Headers for Rows
        add_row_headers(ws1, seating_df)

        # Data for Benches and Seats
        populate_seating_data(ws1, seating_df)

    if not output_filename.endswith(".xlsx"):
        output_filename += ".xlsx"

    try:
        wb.save(output_filename)
        messagebox.showinfo("Success", f"Seating chart saved to {output_filename}")
    except PermissionError:
        messagebox.showerror("Error", f"Permission denied: Cannot save to {output_filename}. Please ensure the file is not open and you have write permissions.")


def add_room_header(ws, room_number, seating_df):
    """Add a room header in the Excel sheet with underlined Room Number and Times New Roman font."""
    # Add Room Number as the header in a new row
    ws.append([f"ROOM {room_number}"])
    
    # Merge cells for the room number header
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=seating_df['Row'].nunique() * 4)
    
    # Center align and set the font for the room number (Times New Roman, size 20, bold, underlined)
    room_header_cell = ws.cell(row=ws.max_row, column=1)
    room_header_cell.alignment = Alignment(horizontal='center')
    
    # Set font to Times New Roman, size 20, underlined, and not bold
    room_header_cell.font = Font(name='Times New Roman', bold=True, size=20, underline='single')

    # Leave one row spacing between the room number and row numbers
    ws.append([''])  # Add an empty row for spacing

    # Call to add row headers immediately after adding the room header
    add_row_headers(ws, seating_df)







def add_row_headers(ws, seating_df):
    """Add headers for each row with Times New Roman font."""
    current_row = ws.max_row + 1  # Increment to place below room header
    for row_num in range(1, seating_df['Row'].nunique() + 1):
        row_header_cell = ws.cell(row=current_row, column=(row_num - 1) * 4 + 1, value=f"Row {row_num}")
        
        # Merge cells for row header
        ws.merge_cells(start_row=current_row, start_column=(row_num - 1) * 4 + 1, end_row=current_row, end_column=(row_num - 1) * 4 + 3)
        
        # Center align and set font to Times New Roman, bold, and size 14
        row_header_cell.alignment = Alignment(horizontal='center')
        row_header_cell.font = Font(name='Times New Roman', bold=True, size=14)





def populate_seating_data(ws, seating_df, left_name, middle_name, right_name):
    """Populate seating data with Times New Roman font and add headers for each bench from the room details."""
    
    # Set bench dimensions (13.57mm width and 50mm length)
    bench_width = 13.57  # in mm
    bench_length = 50  # in mm
    roll_number_font_size = 16  # Set font size for roll numbers

    # Add the header (Names from Excel) above each seat in the bench
    current_row = ws.max_row + 1
    
    for row_num in range(1, seating_df['Row'].nunique() + 1):
        # Insert the corresponding name for each seat position
        for position, name, col_index in zip(["Left", "Middle", "Right"], [left_name, middle_name, right_name], [1, 2, 3]):
            cell = ws.cell(row=current_row, column=(row_num - 1) * 4 + col_index, value=name)
            ws.column_dimensions[cell.column_letter].width = bench_width  # Set width for benches
            
            # Center align and set font to Times New Roman, bold, size 14
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.font = Font(name='Times New Roman', bold=True, size=14)
            cell.border = BORDER_STYLE

    # Move to the next row for actual seating data
    current_row += 1

    # Populate actual seating data below the name headers
    for bench in range(1, seating_df['Bench'].nunique() + 1):
        for seat in range(1, seating_df['Seat'].nunique() + 1):
            for row_num in range(1, seating_df['Row'].nunique() + 1):
                row_df = seating_df[(seating_df['Bench'] == f'Bench {bench}') & (seating_df['Row'] == f'Row {row_num}')]

                # Get seat value and ensure wrapping
                seat_value = row_df[row_df['Seat'] == seat]['Roll Number'].values[0] if not row_df.empty else ''
                col_index = (row_num - 1) * 4 + seat
                seat_cell = ws.cell(row=current_row, column=col_index, value=seat_value)
                
                seat_cell.border = BORDER_STYLE
                seat_cell.font = Font(size=roll_number_font_size)  # Set font size for roll numbers to 16
                seat_cell.alignment = Alignment(horizontal='center', wrap_text=True)  # Enable text wrapping

                # Set row height to accommodate two lines for roll numbers
                ws.row_dimensions[current_row].height = bench_length  # Adjust height for roll numbers
                
        current_row += 1



# Progress window to display the progress bar
def show_progress_window(max_value):
    progress_window = Toplevel()
    progress_window.title("Progress")

    label = Label(progress_window, text="Generating seating chart...")
    label.pack()

    progress_var = StringVar()
    progress_bar = ttk.Progressbar(progress_window, variable=progress_var, maximum=max_value)
    progress_bar.pack()

    return progress_window, progress_var


def main():
    # Create Tkinter root and hide it (used for dialogs)
    root = Tk()
    root.withdraw()

    # Prompt to select room details Excel file
    room_details_path = filedialog.askopenfilename(title="Select Excel file with room details", filetypes=[("Excel files", "*.xlsx *.xls")])
    if not room_details_path:
        messagebox.showerror("Error", "No room details file selected. Exiting...")
        return

    # Load room details DataFrame
    room_details_df = pd.read_excel(room_details_path)
    room_details_df.columns = room_details_df.columns.str.strip()

    # Check required columns in the room details file
    required_columns = ['Room Number', 'Number of Rows', 'Number of Bench', 'Number of Student per Bench', 'Left Path', 'Middle Path', 'Right Path', 'Left Name', 'Middle Name', 'Right Name']
    
    if not all(col in room_details_df.columns for col in required_columns):
        messagebox.showerror("Error", f"Excel file must contain the following columns: {', '.join(required_columns)}.")
        return

    # Get the number of students per bench (from the first row of data)
    students_per_bench = int(room_details_df['Number of Student per Bench'].iloc[0])

    roll_numbers_lists = []
    roll_number_indices = [0] * students_per_bench  # Track indices for each path (Left, Middle, Right)

    # Load roll numbers for Left, Middle, and Right from their respective paths
    for i in range(students_per_bench):
        position = ["Left", "Middle", "Right"][i]
        input_path = room_details_df.iloc[0][f'{position} Path'].strip()
        if not input_path:
            messagebox.showerror("Error", f"No path provided for student position {i + 1}. Exiting...")
            return

        # Load the roll numbers for each position
        df = pd.read_excel(input_path)
        df.columns = df.columns.str.strip()
        if 'Roll Number' not in df.columns:
            messagebox.showerror("Error", f"'Roll Number' column not found in the Excel file for {position}. Exiting...")
            return

        roll_numbers_lists.append(df['Roll Number'].dropna().tolist())

    # Create a progress window for user feedback
    max_value = len(room_details_df.index)
    progress_window, progress_var = show_progress_window(max_value)

    seating_chart_list = []

    # Create a new Excel workbook
    wb = openpyxl.Workbook()

    # Iterate through each room and generate the seating chart
    for room_idx, row in room_details_df.iterrows():
        room_number = row['Room Number']
        rows = int(row['Number of Rows'])
        benches_per_row = int(row['Number of Bench'])

        # Create a new sheet for each room
        sheet_name = f"Room {room_idx + 1}"  # Name the sheets sequentially
        ws = wb.create_sheet(title=sheet_name)

        # Generate seating chart for the current room
        seating_chart_df, roll_number_indices, roll_numbers_exhausted = generate_seating_chart(
            room_number, rows, benches_per_row, students_per_bench, roll_numbers_lists, roll_number_indices, progress_var
        )
        
        seating_chart_list.append(seating_chart_df)

        # Add "Room Number" above the rows, centered
        add_room_header(ws, room_number, seating_chart_df)

        # Extract the names from the row and pass to populate_seating_data
        left_name = row['Left Name']
        middle_name = row['Middle Name']
        right_name = row['Right Name']

        populate_seating_data(ws, seating_chart_df, left_name, middle_name, right_name)

        # If roll numbers are exhausted, prompt for new files
        if roll_numbers_exhausted:
            for i in range(students_per_bench):
                position = SEAT_POSITIONS[i]
                if roll_number_indices[i] >= len(roll_numbers_lists[i]):
                    roll_numbers_lists[i] = prompt_for_new_roll_numbers(position)
                    roll_number_indices[i] = 0

        # Update progress
        progress_var.set(int((len(seating_chart_list) / max_value) * 100))

    # Remove the default sheet created when the workbook was initialized
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Prompt user to select save location for the final Excel file
    output_path = filedialog.asksaveasfilename(title="Save Seating Chart As", filetypes=[("Excel files", "*.xlsx")])
    if not output_path:
        messagebox.showerror("Error", "No file path selected for saving the seating chart. Exiting...")
        return

    # Save the Excel workbook to the specified file
    if not output_path.endswith(".xlsx"):
        output_path += ".xlsx"

    try:
        wb.save(output_path)
        messagebox.showinfo("Success", f"Seating chart saved to {output_path}")
    except PermissionError:
        messagebox.showerror("Error", f"Permission denied: Cannot save to {output_path}. Please ensure the file is not open and you have write permissions.")

    # Close the progress window and root Tkinter window
    progress_window.destroy()
    root.mainloop()



if __name__ == "__main__":
    main()
